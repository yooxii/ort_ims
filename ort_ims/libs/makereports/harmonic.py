"""谐波报告生成"""

import re
from pathlib import Path
from plistlib import InvalidFileException

import openpyxl as xl
from openpyxl.worksheet.worksheet import Worksheet
from tqdm import tqdm
from utils import zips

# 定义常量
MAX_TP_VALUE = 40
ROW_INCREMENT = 19
MAX_ROWS_PER_SECTION = 38
COL_START = 8
COL_INCREMENT = 3
COL_END = 18
DATA_ROWS_COUNT = 19


def _validate_parameters(data_path: str, load_qty: int) -> tuple[int, str]:
    """验证参数有效性"""
    directory = Path(data_path)
    if not directory.exists():
        err = f"文件夹'{directory}'不存在！请重新输入。"
        return 1, err

    if not isinstance(load_qty, int) or load_qty <= 0:
        err = f"负载数量(loadQTY)必须为正整数: {load_qty}"
        return 2, err

    return 0, ""


def _load_template_file(
    template_file: str,
    load_qty: int,
) -> tuple[xl.Workbook | None, str]:
    """加载模板文件"""
    if template_file != "":
        try:
            wb = xl.load_workbook(template_file)
        except OSError as e:
            err = f"打开模版'{template_file}'错误: {e}"
            return None, err
        return wb, ""
    wb = _open_tmplate_file(Path(r".\template"), load_qty)
    if wb is None:
        err = f"未找到'{load_qty}'的模版文件！"
        return None, err
    return wb, ""


def _find_voltage_row(ws: Worksheet, key_row: int, target_voltage: int) -> int | None:
    """查找电压匹配的行"""
    for row in range(key_row, key_row + MAX_ROWS_PER_SECTION, ROW_INCREMENT):
        cell_value = ws.cell(row, 5).value
        if cell_value is not None and int(cell_value) == target_voltage:
            return row
    return None


def _find_load_column(ws: Worksheet, load: str) -> int | None:
    """查找负载匹配的列"""
    for col in range(COL_START, COL_END, COL_INCREMENT):
        cell_value = ws.cell(43, col).value
        if (
            cell_value is not None
            and re.findall(r"\d+", str(cell_value))
            and re.findall(r"\d+", str(cell_value))[0] == load
        ):
            return col
    return None


def _write_data_to_sheet(ws: Worksheet, data, start_row: int, start_col: int):
    """向工作表写入数据"""
    for i in range(DATA_ROWS_COUNT):
        # 检查数据索引是否有效
        if i >= len(data):
            print(f"Warning: Data index {i} out of range")
            continue

        if len(data[i]) < 6:
            print(f"Warning: Data row {i} doesn't have enough columns: {data[i]}")
            continue

        try:
            ws.cell(i + start_row, start_col, float(data[i][5]) * 1000)
            # 设置和下一列合并居中，并且垂直居中
            ws.merge_cells(
                start_row=i + start_row,
                start_column=start_col,
                end_row=i + start_row,
                end_column=start_col + 1,
            )
            ws.cell(i + start_row, start_col).alignment = xl.styles.Alignment(
                horizontal="center", vertical="center"
            )

            # 检查第4列是否有效
            if len(data[i]) < 4:
                print(f"Warning: Data row {i} doesn't have column 4: {data[i]}")
                continue

            ws.cell(i + start_row, start_col + 2, float(data[i][3]) * 1000)
        except (ValueError, IndexError) as e:
            print(f"Error processing data at row {i}: {e}")
            continue


def _process_data_item(ws: Worksheet, data_key: str, data, key_row: int):
    """处理单个数据项"""
    voltage_load_match = re.findall(r"\d+", data_key)
    if len(voltage_load_match) < 2:
        print(f"Warning: Could not extract voltage and load from {data_key}")
        return

    voltage, load = voltage_load_match[0], voltage_load_match[1]
    target_voltage = int(voltage)

    # 查找电压匹配的行
    voltage_row = _find_voltage_row(ws, key_row, target_voltage)
    if voltage_row is None:
        print(f"Warning: Could not find voltage {target_voltage} in the expected range")
        return

    # 查找负载匹配的列
    load_col = _find_load_column(ws, load)
    if load_col is None:
        print(f"Warning: Could not find load {load} in the expected range")
        return

    # 写入数据
    _write_data_to_sheet(ws, data, voltage_row, load_col)


def _is_valid_data_line(line: str):
    """检查数据行是否有效"""
    data = line.split("\t")
    if len(data) == 0 or data[0].strip() == "":
        return False, None

    try:
        tp = int(data[0])
    except ValueError:
        return False, None

    # 只处理奇数次谐波（除了基波1）
    if tp % 2 == 1 and tp != 1:
        return True, data
    if tp == MAX_TP_VALUE:
        return "break", data
    return False, None


def _parse_txt_content(content_lines: list[str]):
    """解析文本内容并返回数据"""
    start_flag = 0
    datas = []

    for line in content_lines:
        if start_flag:
            result, data = _is_valid_data_line(line)
            if result is True:
                datas.append(data)
            elif result == "break":
                break
            # 如果result为False，则跳过该行
            continue

        if start_flag == 0 and re.search(r"\[A\]", line) is not None:
            start_flag = 1

    return datas


def _open_txt(directory: Path):
    """
    从指定目录读取txt文件并解析数据
    """
    if not directory.is_dir():
        return {}

    source_files = [f for f in directory.iterdir() if f.suffix == ".txt"]

    res = {}
    for source_file in source_files:
        try:
            with Path.open(source_file, encoding="utf-8") as fd:
                content_lines = fd.readlines()
                datas = _parse_txt_content(content_lines)
                res[source_file.name] = datas
        except OSError as e:
            print(f"Error reading file {source_file}: {e}")
            continue

    return res


def _deal_datas(directory: Path):
    # 对所有的子目录执行open_txt函数
    res = {}
    for folder in tqdm(Path.iterdir(directory), desc="deal path", ncols=60):
        folder_path = directory / folder
        if not folder_path.is_dir():
            continue
        res[folder] = _open_txt(folder_path)
    return res


def _open_tmplate_file(directory, load_qty):
    """
    打开模版文件

    :param directory: 模版文件路径
    :param load_qty: 负载数量
    :return: xl.workbook对象
    """

    directory = Path(directory)
    template_files = [f for f in directory.iterdir() if f.suffix == "_TP.xlsx"]
    print(template_files)
    for f in template_files:
        spt = f.name.split(" ")
        if load_qty == spt[0]:
            try:
                return xl.load_workbook(f)
            except InvalidFileException as e:
                print(f"Error loading workbook {f}: {e}")
                continue
    return None


def make_harmonic_report(data_path: str, load_qty: int, template_file: str = ""):
    """
    根据所给路径，提取谐波数据并依据模版生成报告

    :param data_path: 数据存放路径
    :param load_qty: 测试负载数量
    :param template_file: 报告模版路径
    :return: 0: 成功, 1: 文件夹不存在, 2: load_qty无效, 3: 模版文件不存在
    """
    _data_path = Path(data_path)
    # 验证参数
    code, err = _validate_parameters(_data_path, load_qty)
    if code != 0:
        print(err)
        return code, err

    # 加载模板文件
    wb, err = _load_template_file(template_file, load_qty)
    if wb is None:
        print(err)
        return 3, err

    # 处理数据
    ws = wb.active
    ws.title = "Harmonics"
    res = _deal_datas(_data_path)
    model = _data_path.name

    # 写入模型名称
    ws["F5"] = model
    key_row = 45

    # 处理每个数据项
    for res_key, datas in tqdm(res.items(), desc="deal datas", ncols=60):
        ws.cell(key_row, 2, res_key)

        for data_key, data in tqdm(datas.items(), desc=f"deal {res_key}", leave=False):
            _process_data_item(ws, data_key, data, key_row)

        key_row += MAX_ROWS_PER_SECTION

    # 保存文件
    save_file = f"{model}_Harmonic Current Test_.xlsx"
    wb.save(save_file)
    zips.zip_folder(_data_path, model)
    return 0, save_file
