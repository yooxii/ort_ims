import win32com.client as win32
import openpyxl as xl
import os
import sys
import json
import shutil
from rapidfuzz import fuzz, process

libs_path = os.path.join(os.path.dirname(__file__), "..")
sys.path.insert(0, libs_path)

from utils.timer import timer

STANDARD_TEMPLATE_DIR = ".\\DataFiles\\report_templates\\# ORT Test Report (WK#)_#"
ORTPLAN_DIR = ".\\DataFiles\\ORTplans"
TEMP_DIR = ".\\DataFiles\\Temp"


def manual_input(info: dict):
    """
    手动输入信息
    """
    res = info.copy()
    for sheet_name in info:
        for title in info[sheet_name]:
            res[sheet_name][title].append(input(title + ": "))

    # TODO
    print(res)
    return res


def open_excel():
    """
    打开Excel应用程序

    Returns:
        excel: Excel应用程序对象
    """

    excel = win32.gencache.EnsureDispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    return excel


def close_excel(excel):
    excel.Quit()


class ORTPlanReport:
    """
    ORT Plan报告生成
    """

    def __init__(self, source_path: str = None, target_path: str = None):
        self.source_path = source_path  # 源文件路径
        self.target_path = target_path  # 目标文件路径
        self.temp_path = TEMP_DIR  # 临时路径
        self.plan_path = ORTPLAN_DIR  # ORT plan路径
        self.template_path = STANDARD_TEMPLATE_DIR  # 模板路径

        self.excel = open_excel()

        self.data = None

    def __delattr__(self, name):
        if self.excel:
            close_excel(self.excel)

    def process_directory(self):
        """处理目录，并将处理完成的文件保存到目标目录下"""
        source_dir = self.source_path
        source_name = source_dir.split("\\")[-1]  # 源文件夹名称
        self.temp_path = temp_report_dir = os.path.join(TEMP_DIR, source_name)

        if not os.path.exists(temp_report_dir):
            os.makedirs(temp_report_dir)

        # 复制源文件夹到临时目录
        shutil.copytree(
            source_dir,
            temp_report_dir,
            ignore=shutil.ignore_patterns("*.json", "*.db"),
            dirs_exist_ok=True,
        )
        # 删除临时目录下的计划模板，为之后生成计划模板做准备
        for filename in os.listdir(temp_report_dir):
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                os.remove(os.path.join(temp_report_dir, filename))
                                                                     
        try:
            # 遍历目标目录
            for filename in os.listdir(source_dir):
                source_file = os.path.abspath(os.path.join(source_dir, filename))

                if filename.endswith(".xls") or filename.endswith(".xlsx"):
                    self.make_plan_template(source_file)
                    break
            shutil.move(temp_report_dir, self.target_dir)
        except Exception as e:
            print(f"process_directory - Error: {e}")
        finally:
            close_excel(self.excel)
            self.excel = None

    @timer
    def copy_ortplan_sheet(self, source_file: str, target_file: str):
        """
        将 source_file 中指定名称的工作表完整复制到 target_file 中。

        Args:
            excel (): Excel 对象
            source_file (str): 源 Excel 文件路径（必须存在）
            target_file (str): 目标 Excel 文件路径（必须存在）
        """
        excel = self.excel
        # 检查文件是否存在
        if not os.path.exists(source_file):
            raise FileNotFoundError(f"源文件不存在: {source_file}")
        if not os.path.exists(target_file):
            raise FileNotFoundError(f"目标文件不存在: {target_file}")

        try:
            # 打开源工作簿和目标工作簿
            source_wb = excel.Workbooks.Open(os.path.abspath(source_file))
            target_wb = excel.Workbooks.Open(os.path.abspath(target_file))

            sheet_names = [s.Name for s in source_wb.Sheets]

            plan_sheetname = ""
            for sheet_name in sheet_names:
                if "plan" in sheet_name.lower():
                    plan_sheetname = sheet_name
                    break
            if plan_sheetname == "":  # 仅处理ORT Plan表
                raise Exception(f"{plan_sheetname}表不存在")

            plan_ws = source_wb.Sheets[plan_sheetname]
            plan_ws.Name = "ORT Plan"
            plan_ws.Copy()

            # 保存ORT Plan表
            new_wb = excel.ActiveWorkbook

            file_name = source_file.split("\\")[-1]
            uut_name = file_name.split(" ")[0]
            date = file_name[file_name.find("WK") : -6]
            save_path = os.path.abspath(
                os.path.join(ORTPLAN_DIR, f"{uut_name}_{date}.xlsx")
            )
            new_wb.SaveAs(save_path, FileFormat=51)
            new_wb.Close()

            # 如果目标文件中已有同名工作表，先删除（避免重命名）
            target_sheetnames = [s.Name for s in target_wb.Sheets]
            for sheet_name in target_sheetnames:
                try:
                    if "plan" in sheet_name.lower():
                        target_wb.Sheets(sheet_name).Delete()
                except:
                    pass  # 不存在则忽略

            # 复制工作表到目标工作簿
            plan_ws.Copy(Before=target_wb.Sheets(2))

            print(
                f"copy_ortplan_sheet - info: 工作表 '{sheet_name}' 已成功复制到 {target_file}"
            )

            # 保存目标工作簿
            target_wb.Save()

        except Exception as e:
            print(f"copy_ortplan_sheet - Error: 发生错误: {e}")
            raise

        finally:
            # 关闭工作簿（不保存源文件）
            if "source_wb" in locals():
                source_wb.Close(SaveChanges=False)
            if "target_wb" in locals():
                target_wb.Close(SaveChanges=False)

        return save_path

    def make_plan_template(self, source_file):
        source_file = os.path.abspath(source_file)
        # 模板文件路径，转为绝对路径
        template_path = os.path.abspath(
            os.path.join(STANDARD_TEMPLATE_DIR, "# ORT Test Report(WK#).xlsx")
        )
        setup_path = os.path.abspath(
            os.path.join(STANDARD_TEMPLATE_DIR, "setup_info.json")
        )

        # 创建模板文件
        temp_file = os.path.abspath(shutil.copy2(template_path, self.temp_path))
        print("make_template - temp_file:" + temp_file)

        # 将源文件中的 ORT Plan 表格复制到模板文件，并保存到ORTPlan目录
        self.ortplan_path = self.copy_ortplan_sheet(source_file, temp_file)
        self.plan_wb = xl.load_workbook(self.ortplan_path, read_only=True)
        self.plan_st = self.plan_wb.active

        # with open(setup_path, "r") as json_file:
        #     info = json.loads(json_file.read())
        # data = manual_input(info)

    def categorize_test_items(self, items: list[str]):
        """
        将测试项目分类，根据TestItems.json中的定义进行匹配

        Args:
            items (list[str]): 测试项目列表

        Returns:
            dict: 包含每个测试项目及其对应类别的字典
        """
        # 读取测试项目分类数据
        with open(
            os.path.join(ORTPLAN_DIR, "TestItems.json"), "r", encoding="utf-8"
        ) as f:
            test_items = json.load(f)

        result = {}

        # 将所有测试项目扁平化，便于查找
        all_tests = {}
        for category, tests in test_items.items():
            for test in tests:
                all_tests[test] = category

        # 遍历每个待分类的测试项目
        for item in items:
            # 使用rapidfuzz查找最佳匹配
            best_match, score, _ = process.extractOne(item, all_tests.keys())

            # 设置阈值，只有分数足够高才认为是有效匹配
            if score > 70:
                result[item] = all_tests[best_match]
            else:
                result[item] = "Uncategorized"

        return result

    @timer
    def get_ortplan_data(self, plan_file: str = None):
        """
        获取ORT Plan数据
        """
        if not plan_file:
            if self.plan_st is None:
                return
            plan_st = self.plan_st
        else:
            plan_st = xl.load_workbook(plan_file).active
        data = []
        data_flag = False
        for row in plan_st.iter_rows(min_row=2, values_only=True):
            if not data_flag:
                # 寻找包含"test items"的列
                col = 0
                for cell in row[:4]:
                    if cell is not None and "items" in str(cell).lower():
                        data_flag = True
                        break
                    col += 1
                if data_flag:
                    continue
            if row[col] is None:
                continue
            if (row[col - 1] is not None) and (row[col + 1] is not None):
                data.append(row[col].strip())
        return data

    def make_teststatus(self, categorized_items: dict = None):
        """
        生成测试状态表

        Args:
            categorized_items (dict): 分类后的测试项

        Returns:
            str: 处理后的文件路径
        """
        data = self.data
        temp_path = self.temp_path
        wb = xl.load_workbook(temp_path)

        data_cover = data["Cover"]
        data_waterfall = data["Waterfall"]
        data_testStatus = data["TestStatus"]

        ########### Cover Sheet ###########
        sheet_cover = wb["Cover"]
        sheet_waterfall = wb["Waterfall"]
        sheet_testStatus = wb["TestStatus for"]

        for package in data_cover.values():
            sheet_cover[package[0]] = package[1]

        ########### Waterfall Sheet ###########
        sn_start_cell = sheet_waterfall[data_waterfall["S/N"][0]]
        sn_start_row = sn_start_cell.row
        sn_start_col = sn_start_cell.column
        if isinstance(data_waterfall["S/N"][1], str):
            sn_str = data_waterfall["S/N"][1].strip()
            if "," in sn_str:
                sn_datas = sn_str.split(",")
            elif ";" in sn_str:
                sn_datas = sn_str.split(";")
            else:
                sn_datas = sn_str.split("")
        else:
            sn_datas = data_waterfall["S/N"][1]
        for i in range(len(sn_datas[1])):
            sheet_waterfall.cell(row=sn_start_row + i, column=sn_start_col).value = (
                sn_datas[i].strip()
            )
        # TODO: 日期的填入

        ########### TestStatus Sheet ###########
        # TODO: TestStatus Sheet

        wb.save(temp_path)


def convert_to_xlsx(excel, file_path):
    """
    将xls文件转换为xlsx格式
    Args:
        excel: Excel应用程序对象
        file_path: 文件路径
    Returns:
        str: 转换后的文件路径
    """
    if file_path.endswith(".xls"):
        try:
            # 打开xls文件
            wb = excel.Workbooks.Open(file_path)
            # 生成新的xlsx文件路径
            xlsx_path = file_path.replace(".xls", ".xlsx")
            # 保存为xlsx格式
            wb.SaveAs(xlsx_path, FileFormat=51)  # 51代表xlsx格式
            wb.Close()
            return xlsx_path
        except Exception as e:
            print(f"convert_to_xlsx - Error: 转换文件时出错: {e}")
            return None
    elif file_path.endswith(".xlsx"):
        return file_path
    else:
        return None


def main():
    """
    主函数：制作模板文件夹
    """
    # 设置源目录和目标目录
    # source_directory = input("请输入源文件夹路径: ")
    # target_directory = input("请输入目标模板文件夹路径: ")
    source_directory = (
        r"DataFiles\reports\FSF050-9TAG ORT Test Report (WK2541)_RT251023"
    )
    target_directory = r"DataFiles\results"

    if not os.path.exists(source_directory):
        print("源文件夹不存在!")
        return
    ort = ORTPlanReport(source_directory,target_directory)
    # 处理目录中的文件
    ort.process_directory()
    print("模板制作完成!")


def test():
    ortplan = ORTPlanReport()
    data = ortplan.get_ortplan_data(r"DataFiles\ORTPlans\FSL028-EL2G_WK254.xlsx")
    print(data)
    cate = ortplan.categorize_test_items(data)
    print(cate)


if __name__ == "__main__":
    main()
    # test()
