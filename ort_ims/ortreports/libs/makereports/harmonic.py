"""谐波报告生成"""

import os
import sys
import re
import openpyxl as xl
from tqdm import tqdm
from pathlib import Path

libs_path = os.path.join(os.path.dirname(__file__), "..")
sys.path.insert(0, libs_path)
from utils.zips import zip_folder


def _open_txt(directory):
    """
    从指定目录读取txt文件并解析数据
    """
    directory = Path(directory)
    if not directory.is_dir():
        return {}

    source_files = [f for f in directory.iterdir() if f.suffix == ".txt"]

    res = {}
    for source_file in source_files:
        try:
            with open(source_file, "r", encoding="utf-8") as fd:
                StartFlag = 0
                datas = []
                for line in fd:
                    if StartFlag:
                        data = line.split("\t")
                        if len(data) == 0 or data[0].strip() == "":
                            continue
                        # tp: 一行数据
                        try:
                            tp = int(data[0])
                        except ValueError:
                            continue  # 跳过无法转换为整数的行

                        if tp % 2 == 1 and tp != 1:
                            datas.append(data)
                        if tp == 40:  # 硬编码值，根据需求保持不变
                            break

                    if StartFlag == 0 and re.search(r"\[A\]", line) is not None:
                        # 找到数据的开始标志
                        StartFlag = 1

                res[source_file.name] = datas
        except IOError as e:
            print(f"Error reading file {source_file}: {e}")
            continue

    return res


def _deal_datas(directory):
    # 对所有的子目录执行open_txt函数
    directory = Path(directory)
    res = {}
    for folder in tqdm(os.listdir(directory), desc="deal path", ncols=60):
        folder_path = directory / folder
        if not folder_path.is_dir():
            continue
        res[folder] = _open_txt(folder_path)
    return res


def _open_tmpFile(directory, loadQTY):
    """
    打开模版文件

    :param directory: 模版文件路径
    :param loadQTY: 负载数量
    :return: xl.workbook对象
    """

    directory = Path(directory)
    TPFiles = [f for f in directory.iterdir() if f.suffix == "_TP.xlsx"]
    print(TPFiles)
    for f in TPFiles:
        spt = f.name.split(" ")
        if loadQTY == spt[0]:
            try:
                return xl.load_workbook(f)
            except Exception as e:
                print(f"Error loading workbook {f}: {e}")
                continue
    return None


def make_harmonic_report(data_path: str, loadQTY: int, tmpFile: str = ""):
    """
    根据所给路径，提取谐波数据并依据模版生成报告

    :param data_path: 数据存放路径
    :param loadQTY: 测试负载数量
    :param tmpFile: 报告模版路径
    :return: 0: 成功, 1: 文件夹不存在, 2: loadQTY无效, 3: 模版文件不存在
    """

    directory = Path(data_path)
    if not directory.exists():
        err = f"文件夹'{directory}'不存在！请重新输入。"
        print(err)
        return 1, err

    # 验证loadQTY是否为有效值
    if not isinstance(loadQTY, int) or loadQTY <= 0:
        err = f"负载数量(loadQTY)必须为正整数: {loadQTY}"
        print(err)
        return 2, err

    if tmpFile != "":
        try:
            wb = xl.load_workbook(tmpFile)
        except OSError as e:
            err = f"打开模版'{tmpFile}'错误: {e}"
            print(err)
            return 3, err
    else:
        wb = _open_tmpFile(Path(r".\template"), loadQTY)
        if wb is None:
            err = f"未找到'{loadQTY}'的模版文件！"
            print(err)
            return 3, err

    ws = wb.active
    ws.title = "Harmonics"
    res = _deal_datas(str(directory))
    model = directory.name

    # 写入数据
    ws["F5"] = model
    key_row = 45
    # 定义常量
    MAX_TP_VALUE = 40
    ROW_INCREMENT = 19
    MAX_ROWS_PER_SECTION = 38
    COL_START = 8
    COL_INCREMENT = 3
    COL_END = 18
    DATA_ROWS_COUNT = 19

    for res_key, datas in tqdm(res.items(), desc="deal datas", ncols=60):
        ws.cell(key_row, 2, res_key)

        for data_key, data in tqdm(datas.items(), desc=f"deal {res_key}", leave=False):
            voltage_load_match = re.findall(r"\d+", data_key)
            if len(voltage_load_match) < 2:
                print(f"Warning: Could not extract voltage and load from {data_key}")
                continue

            voltage, load = voltage_load_match[0], voltage_load_match[1]
            row = key_row
            target_voltage = int(voltage)

            # 查找电压匹配的行
            found_voltage_row = False
            for row in range(key_row, key_row + MAX_ROWS_PER_SECTION, ROW_INCREMENT):
                cell_value = ws.cell(row, 5).value
                if cell_value is not None and int(cell_value) == target_voltage:
                    found_voltage_row = True
                    break

            if not found_voltage_row:
                print(
                    f"Warning: Could not find voltage {target_voltage} in the expected range"
                )
                continue

            # 查找负载匹配的列
            col = COL_START
            found_load_col = False
            for col in range(COL_START, COL_END, COL_INCREMENT):
                cell_value = ws.cell(43, col).value
                if (
                    cell_value is not None
                    and re.findall(r"\d+", str(cell_value))
                    and re.findall(r"\d+", str(cell_value))[0] == load
                ):
                    found_load_col = True
                    break

            if not found_load_col:
                print(f"Warning: Could not find load {load} in the expected range")
                continue

            # 写入数据
            for i in range(0, DATA_ROWS_COUNT):
                # 检查数据索引是否有效
                if i >= len(data):
                    print(f"Warning: Data index {i} out of range for {data_key}")
                    continue

                if len(data[i]) < 6:
                    print(
                        f"Warning: Data row {i} doesn't have enough columns: {data[i]}"
                    )
                    continue

                try:
                    ws.cell(i + row, col, float(data[i][5]) * 1000)
                    # 设置和下一列合并居中，并且垂直居中
                    ws.merge_cells(
                        start_row=i + row,
                        start_column=col,
                        end_row=i + row,
                        end_column=col + 1,
                    )
                    ws.cell(i + row, col).alignment = xl.styles.Alignment(
                        horizontal="center", vertical="center"
                    )

                    # 检查第4列是否有效
                    if len(data[i]) < 4:
                        print(f"Warning: Data row {i} doesn't have column 4: {data[i]}")
                        continue

                    ws.cell(i + row, col + 2, float(data[i][3]) * 1000)
                except (ValueError, IndexError) as e:
                    print(f"Error processing data at row {i}, file {data_key}: {e}")
                    continue
        key_row += MAX_ROWS_PER_SECTION

    # 保存文件
    SaveFile = f"{model}_Harmonic Current Test_.xlsx"
    wb.save(SaveFile)
    zip_folder(directory, model)
    return 0, SaveFile
